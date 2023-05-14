import json
from multiprocessing import process
import os, subprocess, time, win32com, sys, queue, concurrent.futures, csv
from tkinter import font
import platform
import webbrowser
import win32com.client as win32
from tkinter import *
from tkinter import ttk, filedialog, messagebox
from tkinter.ttk import Combobox, Style, Treeview, Notebook, Progressbar
import tkinter as tk, numpy as np, pandas as pd, multiprocessing as mp
import docx2txt, os, glob, re, traceback, fnmatch, PyPDF2, openpyxl, xlrd
from openpyxl import Workbook, load_workbook
from threading import Thread
from pathlib import Path
from ttkthemes import ThemedStyle
from pandastable import Table, TableModel
from fuzzywuzzy import fuzz


# Legg til CSV
# Fix første kolonne i treeview for source så den kan utvides.

#TODO: Finn all words ikke riktig etter cancel

class MainWindow:
    # define constructor, make everything inside stretch when the window is resized
    def __init__(self, master):
        self.master = master
        self.Source_file_name = ''
        self.Target_file_name = ''

        self.master.bind_all("<Escape>", self.cancel_button_pressed)

        main_window = tk.Frame(self.master)
        main_window.pack(fill=BOTH, expand=True)
        self.Search_frame = ttk.Frame(self.master)
        # Initialize the search style
        self.Search_style = ttk.Style()
        try:
            self.load_settings_for_program()
        except:
            pass
        #self.Search_style.theme_use('vista')

        # create a notebook widget and add it to main_window
        self.notebook_main = ttk.Notebook(main_window)
        self.notebook_main.pack(fill=BOTH, expand=True)

        self.notebook_main.grid_rowconfigure(0, weight=1)
        # Use the uniform option to distribute extra space equally between the two frames
        self.notebook_main.grid_rowconfigure(0, weight=1, uniform="a")
        self.notebook_main.grid_rowconfigure(1, weight=1, uniform="a")

        # create the pages
        self.search_and_index_page = ttk.Frame(self.notebook_main)
        self.search_in_folder_page = ttk.Frame(self.notebook_main)

        # add the pages to the notebook
        self.notebook_main.add(self.search_and_index_page, text="Search and Index")
        self.notebook_main.add(self.search_in_folder_page, text="Search in folder")

        self.top_frame = Frame(self.search_and_index_page, height=50)
        self.top_frame.pack(side=TOP, fill=BOTH, expand=False)
        self.top_frame.pack_propagate(False)

        self.bottom_frame = Frame(self.search_and_index_page)
        self.bottom_frame.pack(side=BOTTOM, fill=BOTH, expand=True)
        
        master.rowconfigure(0, weight=1)
        master.columnconfigure(0, weight=1)
        main_window.columnconfigure(0, weight=1)
        main_window.rowconfigure(1, weight=1)

        # create labels
        self.Source_label = tk.Label(self.top_frame, text="Source file:") 
        self.Source_label.grid(row=0, column=0, sticky=W, padx=(10, 0), pady=(0, 5))

        self.Target_label = tk.Label(self.top_frame, text="Target file:") 
        self.Target_label.grid(row=1, column=0, sticky=W, padx=(10, 0), pady=(0, 5))

        self.Source_file_name_label = tk.Label(self.top_frame, text="Choose file 1:") 
        # stick to left, dont stretch, add 10 pixels to the left and 10 pixels to the top
        self.Source_file_name_label.grid(row=0, column=1, sticky=W, padx=(10, 0), pady=(0, 5))

        self.Target_file_name_label = tk.Label(self.top_frame, text="Choose file 2:") 
        self.Target_file_name_label.grid(row=1, column=1, sticky=W, padx=(10, 0), pady=(0, 5))

        self.Source_columns_label = tk.Label(self.top_frame, text="Select source column:") 
        self.Source_columns_label.grid(row=2, column=0, sticky=W, padx=(10, 0), pady=(0, 5))

        self.Target_columns_label = tk.Label(self.top_frame, text="Select target column:") 
        self.Target_columns_label.grid(row=3, column=0, sticky=W, padx=(10, 0), pady=(0, 5))

        # # create dropdown Source_column_combo, add after Select source column: and before Browse
        self.Source_column_combo = ttk.Combobox(self.top_frame, width=100, state='readonly')
        self.Source_column_combo.grid(row=2, column=1, sticky=W, padx=(0, 10), pady=(0, 5))

        # # create dropdown Target_column_combo, add after Select target column: and before Browse
        self.Target_column_combo = ttk.Combobox(self.top_frame, width=100, state='readonly')
        self.Target_column_combo.grid(row=3, column=1, sticky=W, padx=(0, 10), pady=(0, 5))

        self.Source_column_combo.config(state='disabled')
        self.Target_column_combo.config(state='disabled')

        # create buttons
        self.Source_browse_button = tk.Button(self.top_frame, text="Browse", width=12,
                                              command=self.get_filename1) 
        self.Source_browse_button.grid(row=0, column=3, sticky=E, padx=(0, 10), pady=(0, 5))

        self.Target_browse_button = tk.Button(self.top_frame, text="Browse", width=12,
                                              command=self.get_filename2) 
        self.Target_browse_button.grid(row=1, column=3, sticky=E, padx=(0, 10), pady=(0, 5))

        self.Compare_files_button = tk.Button(self.top_frame, text="Compare", width=12,
                                              command=self.compare_files) 
        self.Compare_files_button.grid(row=4, column=0, sticky=W, padx=(0, 10), pady=(0, 5))
        self.Compare_files_button.config(state=tk.DISABLED)

        self.top_frame.columnconfigure(6, weight=1) # Configure column 6 to expand

        button_font = font.Font(size=15)
        self.help_button = Button(self.top_frame, text="🤔", font=button_font, command=self.help_button)
        self.help_button.grid(row=0, column=6, sticky=E, padx=(0, 10), pady=(0, 5)) # Change padx to (0, 10) to have spacing only on the right side

        self.theme_button = Button(self.top_frame, text="🎨", font=button_font, command=self.change_theme)
        self.theme_button.grid(row=1, column=6, sticky=E, padx=(0, 10), pady=(0, 5)) # Change padx to (0, 10) to have spacing only on the right side


        self.split_text_button = tk.Button(self.top_frame, text="Split text", width=12,
                                      command=self.split_text) 
        self.split_text_button.grid(row=4, column=1, sticky=W, padx=(0, 0), pady=(0, 5))
        #disable
        self.split_text_button.config(state=tk.DISABLED)

        self.find_word_button = tk.Button(self.top_frame, text="Find all words", width=12,
                                      command=self.find_all_words) 
        # stick to split text button, dont stretch, add 10 pixels to the left and 10 pixels to the top   
        self.find_word_button.grid(row=4, column=1, sticky=W, padx=(120, 0), pady=(0, 5))
        #disable
        self.find_word_button.config(state=tk.DISABLED)

        self.Clear_button = tk.Button(self.top_frame, text="Clear", width=12,
                                      command=self.clear) 
        self.Clear_button.grid(row=4, column=3, sticky=W, padx=(0, 0), pady=(0, 5))
        
        #add self.notebook to bottom frame, stretch to fill frame
        self.notebook = ttk.Notebook(self.bottom_frame)
        self.notebook.pack(fill=BOTH, expand=True)

        # Configure the rows and columns of the notebook to expand and fill the available space
        self.notebook.grid_rowconfigure(0, weight=1)
        self.notebook.grid_columnconfigure(0, weight=1)
        
        # Create a frame and add it to the main window
        self.Input_files_frame = ttk.Frame(self.notebook)
        self.Input_files_frame.grid(row=0, column=0, sticky="nsew")
        self.Input_files_frame.grid_rowconfigure(0, weight=1)
        # Use the uniform option to distribute extra space equally between the two frames
        self.Input_files_frame.grid_columnconfigure(0, weight=1, uniform="a")
        self.Input_files_frame.grid_columnconfigure(1, weight=1, uniform="a")

        self.notebook.add(self.Input_files_frame, text="Input files")

        # middle mouse click on tabs in notebook to close tab, but not the first tab
        self.notebook.bind("<Button-2>", self.close_tab)

        self.Input_left_frame = ttk.Frame(self.Input_files_frame)
        self.Input_left_frame.grid(row=0, column=0, sticky="nsew")
        self.Input_left_frame.grid_rowconfigure(0, weight=1)

        self.Input_right_frame = ttk.Frame(self.Input_files_frame)
        self.Input_right_frame.grid(row=0, column=1, sticky="nsew")
        self.Input_right_frame.grid_rowconfigure(0, weight=1)

        self.Input_left_frame.grid_rowconfigure(0, weight=1)
        self.Input_left_frame.grid_columnconfigure(0, weight=1)
        self.Input_right_frame.grid_rowconfigure(0, weight=1)
        self.Input_right_frame.grid_columnconfigure(0, weight=1)

        #Configure the Source_treeview and its scrollbar to fill the Input_left_frame
        self.Source_treeview = Treeview(self.Input_left_frame)
        self.Source_treeview_scrollbar_vertical = ttk.Scrollbar(self.Input_left_frame, orient="vertical", command=self.Source_treeview.yview)
        self.Source_treeview_scrollbar_horizontal = ttk.Scrollbar(self.Input_left_frame, orient="horizontal", command=self.Source_treeview.xview)
        self.Source_treeview.configure(yscrollcommand=self.Source_treeview_scrollbar_vertical.set, xscrollcommand=self.Source_treeview_scrollbar_horizontal.set)

        self.Source_treeview.grid(row=0, column=0, sticky="nsew")
        self.Source_treeview.grid_rowconfigure(0, weight=1)
        self.Source_treeview.grid_columnconfigure(0, weight=1)
        self.Source_treeview_scrollbar_vertical.grid(row=0, column=1, sticky="nswe")
        self.Source_treeview_scrollbar_horizontal.grid(row=1, column=0, sticky="ew")

        self.Target_treeview = Treeview(self.Input_right_frame)
        self.Target_treeview_scrollbar_vertical = ttk.Scrollbar(self.Input_right_frame, orient="vertical", command=self.Target_treeview.yview)
        self.Target_treeview_scrollbar_horizontal = ttk.Scrollbar(self.Input_right_frame, orient="horizontal", command=self.Target_treeview.xview)
        self.Target_treeview.configure(yscrollcommand=self.Target_treeview_scrollbar_vertical.set, xscrollcommand=self.Target_treeview_scrollbar_horizontal.set)
        

        self.Target_treeview.grid(row=0, column=0, sticky="nsew")
        self.Target_treeview.grid_rowconfigure(0, weight=1)
        self.Target_treeview.grid_columnconfigure(0, weight=1)
        self.Target_treeview_scrollbar_vertical.grid(row=0, column=1, sticky="nswe")
        self.Target_treeview_scrollbar_horizontal.grid(row=1, column=0, sticky="ew")

        self.Input_left_frame.grid_rowconfigure(0, weight=1)
        self.Input_left_frame.grid_columnconfigure(0, weight=1)
        self.Input_right_frame.grid_rowconfigure(0, weight=1)
        self.Input_right_frame.grid_columnconfigure(0, weight=1)

        
        ##############################################################################################
        #######################################  SEARCH IN GUI #######################################
        ##############################################################################################
        # self.Search_style = Style()
        # self.Search_style.theme_use("default")

        Search_wordinput = Frame(self.search_in_folder_page)
        Search_wordinput.pack(fill=X)

        lbl1 = Label(Search_wordinput, text="Search for:", width=10)
        lbl1.grid(row=0, column=0, padx=5, pady=5)



        self.Search_entry1 = Entry(Search_wordinput, width=50)
        self.Search_entry1.grid(row=0, column=1, padx=5, pady=5)

        self.Search_button1 = Button(Search_wordinput, text="Select folder", command=self.select_folder)
        self.Search_button1.grid(row=0, column=2, padx=5, pady=5)

        Search_select_folder = Frame(self.search_in_folder_page)
        Search_select_folder.pack(fill=X)

        self.label_select_folder = Label(Search_select_folder, text="No folder selected", bg='#FFCCCB', foreground="black")
        self.label_select_folder.pack(side=LEFT, padx=5, expand=False)

        frame2 = Frame(self.search_in_folder_page)
        frame2.pack(fill=X)

        self.Search_box_value = StringVar()
        self.Search_box = Combobox(frame2, textvariable=self.Search_box_value, state='readonly')
        self.Search_box['values'] = ('Excel','CSV', 'Word', 'PDF', 'Excel, CSV, Word and PDF')
        self.Search_box.current(0)
        self.Search_box.pack(side=LEFT, padx=5, pady=5)

        self.Search_box2_value = StringVar()
        self.Search_box2 = Combobox(frame2, textvariable=self.Search_box2_value, state='readonly')
        self.Search_box2['values'] = ('Exact match', 'Partial match', "fuzzy_match")
        self.Search_box2.current(0)
        self.Search_box2.pack(side=LEFT, padx=5, pady=5)

        frame3 = Frame(self.search_in_folder_page)
        frame3.pack(fill=X, padx=5, pady=5)

        # use mp.Process to run the search function
        self.Search_btn1 = Button(frame3, text="Search", command=self.Search_OnButtonClick)
        self.Search_btn1.pack(side=LEFT, padx=5, pady=5)
        #disable self.Search_btn1
        self.Search_btn1.config(state='disabled')

        self.Search_btn3 = Button(frame3, text="Clear", command=self.Search_Clear)
        self.Search_btn3.pack(side=LEFT, padx=5, pady=5)
            # create a larger font for the button text
        button_font = font.Font(size=12)
        self.Search_btn4 = Button(frame3, text="\U0001F4BE",font=button_font, command=self.Search_Save)
        self.Search_btn4.pack(side=LEFT, padx=5, pady=5)
        
        # if checkbutton is pressed search for in all subfolder in selected folder
        self.Search_checkbutton_value = IntVar()
        self.Search_checkbutton = Checkbutton(frame3, text="Search in subfolders", variable=self.Search_checkbutton_value)
        self.Search_checkbutton.pack(side=LEFT, padx=5, pady=5)

        # # mouseover tooltip on checkbutton "Search in subfolders might take a while"
        self.Search_checkbutton.bind("<Enter>", self.Search_OnEnter)
        self.Search_checkbutton.bind("<Leave>", self.Search_OnLeave)

        self.Search_progress = Progressbar(self.search_in_folder_page, orient=HORIZONTAL, length=100, mode='determinate')
        self.Search_progress.pack(fill=X, padx=5, pady=5)

        self.Search_status = tk.Label(self.search_in_folder_page, text="Please select a folder to search.", bd=1, relief=SUNKEN, anchor=W)
        self.Search_status.pack(fill=X, padx=5, pady=5, side=BOTTOM)

# set presize size on column, auto size when data is present
        self.Search_tree = Treeview(self.search_in_folder_page, columns=("", "File", "Sheet", "Cell", "Value"))
        self.Search_tree.heading('#0', text='', anchor=W)
        self.Search_tree.heading('#1', text='File', anchor=W)
        self.Search_tree.heading('#2', text='Sheet', anchor=W)
        self.Search_tree.heading('#3', text='Cell', anchor=W)
        self.Search_tree.heading('#4', text='Value', anchor=W)
        self.Search_tree.column('#0', stretch=NO, minwidth=10, width=0)
        self.Search_tree.column('#1', stretch=YES, minwidth=300, width=380)
        self.Search_tree.column('#2', stretch=YES, minwidth=50, width=100)
        self.Search_tree.column('#3', stretch=YES, minwidth=30, width=50)
        self.Search_tree.column('#4', stretch=YES, minwidth=500, width=800)

        # scroll bar for treeview horizontally and vertically
        self.Search_scrollbar = ttk.Scrollbar(self.search_in_folder_page, orient="vertical", command=self.Search_tree.yview)
        self.Search_scrollbar.pack(side=RIGHT, fill=Y)
        self.Search_scrollbar2 = ttk.Scrollbar(self.search_in_folder_page, orient="horizontal", command=self.Search_tree.xview)
        self.Search_scrollbar2.pack(side=BOTTOM, fill=X)
        self.Search_tree.configure(yscrollcommand=self.Search_scrollbar.set)
        self.Search_tree.pack(fill=BOTH, expand=True, padx=5, pady=5)
        self.Search_tree.bind("<Double-1>", self.Search_OnDoubleClick)
        #self.Search_tree.bind("<Enter>", self.Search_hover)
        # on sigle click
        self.Search_tree.bind("<Button-3>", self.Search_single_click)


        # create file dialog to choose file from user
    def get_filename1(self):      
        # if cancel is pressed, do nothing
        self.Source_file_name = filedialog.askopenfilename(initialdir=os.getcwd(),
                                            title="Select file1",
                                            filetypes=(("xlsx files", "*.xlsx"),
                                                        ("csv files", "*.csv"),
                                                        ("all files", "*.*")))

        self.split_text_button.config(state=tk.NORMAL)
        # if file is csv, read csv file and create dataframe and add to dropdown menu
        if self.Source_file_name.endswith('.csv'):
            df1 = pd.read_csv(self.Source_file_name)
            columns1 = df1.columns.tolist()
            self.Source_column_combo['values'] = columns1
            self.Source_column_combo.current(0)
            self.Source_column_combo.config(state='readonly')
        # if file is excel, read excel file and create dataframe and add to dropdown menu
        elif self.Source_file_name.endswith('.xlsx'):
            df1 = pd.read_excel(self.Source_file_name, sheet_name=0, engine='openpyxl', index_col=None, na_values=['NA'])
            columns1 = df1.columns.tolist()
            self.Source_column_combo['values'] = columns1
            self.Source_column_combo.current(0)
            self.Source_column_combo.config(state='readonly')
        else:
            pass
            #self.split_text_button.config(state=tk.NORMAL)

        if self.Source_file_name != '':
            self.Source_file_name_label.config(text=self.Source_file_name)
            self.df1 = df1
            self.populate_treeview()
        else:
            pass
            #self.Source_file_name_label.configure(text='')
            #self.Source_file_name_label.config(text='Choose file 1:')

        self.populate_treeview()

        # endable compare button if both files are selected
        if self.Source_file_name != '' and self.Target_file_name != '':
            self.Compare_files_button.config(state=tk.NORMAL)
            self.find_word_button.config(state=tk.NORMAL)
            


    # check if file 2 is selected, and vice versa enable self.mergefromfile2_button.config(state=NORMAL) if both is selected
    def get_filename2(self):      
        # if cancel is pressed, do nothing
        self.Target_file_name = filedialog.askopenfilename(initialdir=os.getcwd(),
                                            title="Select file1",
                                            filetypes=(("xlsx files", "*.xlsx"),
                                                        ("csv files", "*.csv"),
                                                        ("all files", "*.*")))

        # if file is csv, read csv file and create dataframe and add to dropdown menu
        if self.Target_file_name.endswith('.csv'):
            df2 = pd.read_csv(self.Target_file_name)
            columns2 = df2.columns.tolist()
            self.Target_column_combo['values'] = columns2
            self.Target_column_combo.current(0)
            self.Target_column_combo.config(state='readonly')
        # if file is excel, read excel file and create dataframe and add to dropdown menu
        elif self.Target_file_name.endswith('.xlsx'):
            df2 = pd.read_excel(self.Target_file_name, sheet_name=0, engine='openpyxl', index_col=None, na_values=['NA'])
            columns2 = df2.columns.tolist()
            self.Target_column_combo['values'] = columns2
            self.Target_column_combo.current(0)
            self.Target_column_combo.config(state='readonly')
        else:
            pass
            #self.split_text_button.config(state=tk.NORMAL)

        if self.Target_file_name != '':
            self.Target_file_name_label.config(text=self.Target_file_name)
            self.df2 = df2
            self.populate_treeview()
        else:
            pass

        self.populate_treeview()

        # endable compare button if both files are selected
        if self.Source_file_name != '' and self.Target_file_name != '':
            self.Compare_files_button.config(state=tk.NORMAL)
            self.find_word_button.config(state=tk.NORMAL)

    def populate_treeview(self):
            
        if self.Source_file_name:
            # load source file into pandas dataframe
            # if excel file
            if self.Source_file_name.endswith('.xlsx'):
                self.Source_dataframe = pd.read_excel(self.Source_file_name)

            # if csv file
            try:
                if self.Source_file_name.endswith('.csv'):
                    self.Source_dataframe = pd.read_csv(self.Source_file_name)
            except:
                # if error show message "Error reading: Try converting to excel file"
                messagebox.showerror("Error", "Error reading file: Try converting to excel file")
                return

            # remove nan values from dataframe
            self.Source_dataframe = self.Source_dataframe.fillna('')
            # clear the treeview
            self.Source_treeview.delete(*self.Source_treeview.get_children())

            # hide first column in treeview (index column), and set rest to atleast 100 width for all columns in treeview
            self.Source_treeview.column('#0', width=0, stretch=False)
            # add header name to the treeview from column in self.Target_dataframe, drop the first start at second header name with stretch=tk.YES
            self.Source_treeview['column'] = list(self.Source_dataframe.columns.values)[0:]
            
            # set width for all columns in treeview
            for column in list(self.Source_dataframe.columns.values)[0:]:
                self.Source_treeview.column(column, minwidth=150, width=150, stretch=True)

            # add header name to the treeview from column in self.Target_dataframe, drop the first start at second header name with stretch=tk.YES
            for column in self.Source_treeview['columns']:
                self.Source_treeview.heading(column, text=column)
                    
            # add column values to the treeview
            for row in self.Source_dataframe.itertuples(index=True, name='Pandas'):
                self.Source_treeview.insert('', 'end', values=row[1:])
                


        if self.Target_file_name:
            #if excel file
            if self.Target_file_name.endswith('.xlsx'):
                self.Target_dataframe = pd.read_excel(self.Target_file_name)
            #if csv file
            # if csv file
            try:
                if self.Target_file_name.endswith('.csv'):
                    self.Target_dataframe = pd.read_csv(self.Target_file_name)
            except:
                # if error show message "Error reading: Try converting to excel file"
                messagebox.showerror("Error", "Error reading file: Try converting to excel file")
                return
            
        # remove nan values from dataframe
            self.Target_dataframe = self.Target_dataframe.fillna('')
            # clear the treeview
            self.Target_treeview.delete(*self.Target_treeview.get_children())
            # hide first column in treeview (index column)
            self.Target_treeview.column('#0', width=0, stretch=False)
            # add header name to the treeview from column in self.Target_dataframe, drop the first start at second header name
            self.Target_treeview['columns'] = list(self.Target_dataframe.columns.values)[0:]
            # set rest to atleast 100 width for all columns in treeview, but the index column is 0
            for column in list(self.Target_dataframe.columns.values)[0:]:
                self.Target_treeview.column(column, minwidth=100, width=100, stretch=True)


            # add column header to the treeview
            for column in self.Target_treeview['columns']:
                self.Target_treeview.heading(column, text=column)
            
            for row in self.Target_dataframe.itertuples(index=True, name='Pandas'):
                self.Target_treeview.insert('', 'end', values=row[1:])
                
        # select page "Compare files" in self.notebook
        self.notebook.select(self.Input_files_frame)

    def compare_files(self):

        # if page "Compare files" already exist, delete it and create new one
        try:
            if self.notebook.tab(self.Compare_files_frame, option="text") == "Compare files":
                # get index of tab "Compare files"
                #tab_number = self.notebook.index(self.Compare_files_frame)
                self.notebook.forget(self.Compare_files_frame)
                
                self.Compare_files_frame = Frame(self.notebook)
                self.Compare_files_frame.pack(fill=BOTH, expand=True)
                self.notebook.add(self.Compare_files_frame, text="Compare files")
                self.notebook.pack(fill="both", expand=True)
                # activate compare files frame in self.notebook
                self.notebook.select(self.Compare_files_frame)
                self.Compare_files_frame_left = tk.Frame(self.Compare_files_frame)
                self.Compare_files_frame_right = tk.Frame(self.Compare_files_frame)
                self.Compare_files_frame_left.grid(row=0, column=0, sticky='nsew')
                self.Compare_files_frame_right.grid(row=0, column=1, sticky='nsew')
                self.Compare_files_frame_left.columnconfigure(0, weight=1)
                self.Compare_files_frame_left.rowconfigure(1, weight=1)
                self.Compare_files_frame_right.columnconfigure(0, weight=1)
                self.Compare_files_frame_right.rowconfigure(1, weight=1)
               
        except:
            self.Compare_files_frame = Frame(self.notebook)
            self.Compare_files_frame.pack(fill=BOTH, expand=True)
            self.notebook.add(self.Compare_files_frame, text="Compare files")
            self.notebook.pack(fill="both", expand=True)
            # activate compare files frame in self.notebook
            self.notebook.select(self.Compare_files_frame)
            # make 2 frames in Compare_files_frame Compare_files_frame_left and Compare_files_frame_right
            self.Compare_files_frame_left = tk.Frame(self.Compare_files_frame)
            self.Compare_files_frame_right = tk.Frame(self.Compare_files_frame)
            self.Compare_files_frame_left.grid(row=0, column=0, sticky='nsew')
            self.Compare_files_frame_left.columnconfigure(0, weight=1)
            self.Compare_files_frame_left.rowconfigure(1, weight=1)
            self.Compare_files_frame_right.grid(row=0, column=1, sticky='nsew')
            self.Compare_files_frame_right.columnconfigure(0, weight=1)
            self.Compare_files_frame_right.rowconfigure(1, weight=1)


        self.column1 = self.Source_column_combo.get()
        self.column2 = self.Target_column_combo.get()

        # if int convert to string
        if type(self.column1) == int:
            self.column1 = str(self.column1)
        if type(self.column2) == int:
            self.column2 = str(self.column2)

        self.df1_column = self.df1[self.column1]
        self.df2_column = self.df2[self.column2]

        self.equal_values = []
        self.unequal_values = []

        for i in range(len(self.df1_column)):
            if self.df1_column[i] in self.df2_column.values:
                self.equal_values.append(self.df1_column[i])
            else:
                self.unequal_values.append(self.df1_column[i])


        # create scrollbar vertical
        self.scrollbar1 = tk.Scrollbar(self.Compare_files_frame_left, orient='vertical')
        self.scrollbar2 = tk.Scrollbar(self.Compare_files_frame_right, orient='vertical')
        
        # create scrollbar horizontal
        self.scrollbar3 = tk.Scrollbar(self.Compare_files_frame_left, orient='horizontal')
        self.scrollbar4 = tk.Scrollbar(self.Compare_files_frame_right, orient='horizontal')

        # create label on top
        self.Equal_label = tk.Label(self.Compare_files_frame_left, text="Equal Values")
        self.Unequal_label = tk.Label(self.Compare_files_frame_right, text="Unequal Values")
       
        # place label on top 
        self.Equal_listbox = tk.Listbox(self.Compare_files_frame_left, yscrollcommand=self.scrollbar1.set, xscrollcommand=self.scrollbar3.set, width=50, height=20)
        self.Unequal_listbox = tk.Listbox(self.Compare_files_frame_right, yscrollcommand=self.scrollbar2.set, xscrollcommand=self.scrollbar4.set, width=50, height=20)
        
        # place scrollbar
        self.scrollbar1.grid(row=1, column=1, sticky='ns')
        self.scrollbar2.grid(row=1, column=1, sticky='ns')
        self.scrollbar3.grid(row=2, column=0, sticky='ew')
        self.scrollbar4.grid(row=2, column=0, sticky='ew')

        # place listbox
        self.Equal_listbox.grid(row=1, column=0, sticky='nsew')
        self.Unequal_listbox.grid(row=1, column=0, sticky='nsew')
        
        # place label
        self.Equal_label.grid(row=0, column=0, sticky='nsew')
        self.Unequal_label.grid(row=0, column=0, sticky='nsew')
                
        # insert equal values in listbox
        for value in self.equal_values:
            self.Equal_listbox.insert(tk.END, value)
        
        # insert unequal values in listbox
        for value in self.unequal_values:
            self.Unequal_listbox.insert(tk.END, value)
        
        # configure window frame
        self.Compare_files_frame.grid_rowconfigure(0, weight=1)
        self.Compare_files_frame.grid_columnconfigure(0, weight=1)
        self.Compare_files_frame.grid_columnconfigure(1, weight=1)
        
        # configure scrollbar
        self.scrollbar1.config(command=self.Equal_listbox.yview)
        self.scrollbar2.config(command=self.Unequal_listbox.yview)
        self.scrollbar3.config(command=self.Equal_listbox.xview)
        self.scrollbar4.config(command=self.Unequal_listbox.xview)       

        # Create search text input
        self.search_text = tk.StringVar()
        self.search_text_entry = tk.Entry(self.Compare_files_frame, textvariable=self.search_text)
        self.search_text_entry.grid(row=3, column=0, sticky='nsew')
        self.search_text_entry.bind('<Return>', self.search_in_listbox)

        # Create search button
        #self.search_button = tk.Button(self.Compare_files_frame, text="Search", command=self.search_in_listbox)
        #self.search_text_entry.bind('<Return>', self.search_in_listbox)

        # add button Output_format to frame Compare_files_frame, make exstantion when button is pressed add show extra frame to the right of Compare_files_frame and add 2 listboxes
        self.Output_format_button = tk.Button(self.Compare_files_frame, text="Output format", command=self.Mergecolumns)
        self.Output_format_button.grid(row=3, column=1, sticky='nsew')

    def help_button(self):
        # open Assets/SearchandIndexHelp.pdf
        os.startfile(r"Ny\SearchandIndex\Assets\SearchandIndexHelp.pdf")

    def change_theme(self):
        # Set of available themes
        THEMES = ["radiance", "clam", "alt", "default", "classic", "vista", "aquativo"]
        # Get current theme index
        current_theme_index = THEMES.index(self.Search_style.theme_use())
        # Get index of next theme in the list of available themes
        next_theme_index = (current_theme_index + 1) % len(THEMES)
        # Get name of next theme
        next_theme = THEMES[next_theme_index]
        # Change theme to next theme
        if next_theme in ["adapta", "radiance", "aquativo"]:
            self.Search_style = ThemedStyle(self.Search_frame)
            self.Search_style.set_theme(next_theme)
        else:
            self.Search_style.theme_use(next_theme)
        # Update the "theme" setting and save to file
        self.settings["theme"] = next_theme
        self.save_settings_for_program()

    def save_settings_for_program(self):
        # Save settings to file
        with open("settings.json", "w") as f:
            json.dump(self.settings, f)

    def load_settings_for_program(self):
        # Load settings from file
        try:
            with open("settings.json", "r") as f:
                self.settings = json.load(f)
        except FileNotFoundError:
            # If file does not exist, create a default settings dictionary
            self.settings = {"theme": "radiance", "setting1": True, "setting2": "some_value"}
        # Set the theme to the last used theme
        self.set_theme(self.settings["theme"])

    def update_settings(self, setting_name, value):
        # Update a specific setting and save to file
        self.settings[setting_name] = value
        self.save_settings_for_program()

    def set_theme(self, theme):
        # Set the theme and update settings
        self.Search_style.theme_use(theme)
        self.update_settings("theme", theme)


    def update_settings(self, setting_name, value):
        # Update a specific setting and save to file
        self.settings[setting_name] = value
        self.save_settings_for_program()



    # for self.equal_values find in file2 and merge columns PS and Tag for file1
    def Mergecolumns(self):
        # if page "Output format frame" already exist, delete it and create new one
        try:
            if self.notebook.tab(self.Output_format_frame, option="text") == "Output format":
                self.notebook.forget(self.Output_format_frame)
                self.Output_format_frame = ttk.Frame(self.notebook)
                self.notebook.add(self.Output_format_frame, text='Output format')
                #navigate to tab "Output format"
                self.notebook.select(self.Output_format_frame)        
                self.search_in_listbox = MainWindow.search_in_listbox

        except:
            # create new notebook with tabname "Output format" in mainwindow
            self.Output_format_frame = ttk.Frame(self.notebook)
            self.notebook.add(self.Output_format_frame, text='Output format')
            #navigate to tab "Output format"
            self.notebook.select(self.Output_format_frame)        
            self.search_in_listbox = MainWindow.search_in_listbox

        # inside Output:_format_frame create 2 frames, on the left side the listbox with all columns from file1 and file2
        self.Output_left_frame = ttk.Frame(self.Output_format_frame, width=200)
        self.Output_left_frame.pack(fill='both', expand=False, side='left', padx=(10, 0), pady=(10, 0))

        self.Output_right_frame = ttk.Frame(self.Output_format_frame)
        self.Output_right_frame.pack(fill='both', expand=True, side='left', padx=(10, 0), pady=(10, 0))

        self.Output_treeview = ttk.Frame(self.Output_right_frame)
        self.Output_treeview.pack(fill='both', expand=True)

        self.Output_treeview = ttk.Treeview(self.Output_treeview)
        self.Output_treeview.pack(fill='both', expand=True)

        self.scrollbar_treeview_vertical = tk.Scrollbar(self.Output_treeview, orient="vertical", command=self.Output_treeview.yview)
        self.scrollbar_treeview_vertical.pack(side=RIGHT, fill=Y)

        self.scrollbar_treeview_Horizontal = tk.Scrollbar(self.Output_treeview, orient="horizontal", command=self.Output_treeview.xview)
        self.scrollbar_treeview_Horizontal.pack(side=BOTTOM, fill=X)

        self.Output_treeview.configure(yscrollcommand=self.scrollbar_treeview_vertical.set, xscrollcommand=self.scrollbar_treeview_Horizontal.set)
        self.Output_treeview.configure(xscrollcommand=self.scrollbar_treeview_Horizontal.set)


        # errorhandling if no file is selected
        if self.Source_file_name == "" or self.Target_file_name == "":
            messagebox.showinfo("Error", "Please select files and run 'Compare' first")
        else:
            self.df_merged = pd.merge(self.df1, self.df2, how='left', left_on=self.column1, right_on=self.column2)

            # Create two new frames inside the Output_left_frame
            self.left_top_frame = ttk.Frame(self.Output_left_frame)
            self.left_top_frame.pack(fill='both', expand=True, side='top', padx=(10, 0), pady=(10, 0))

            self.left_bottom_frame = ttk.Frame(self.Output_left_frame)
            self.left_bottom_frame.pack(fill='both', expand=False, side='bottom', padx=(10, 0), pady=(10, 0))

            # Create the widgets in the left_top_frame
            self.listbox1 = tk.Listbox(self.left_top_frame, selectmode='multiple')
            self.listbox1.pack(fill='both', expand=True, side='left', padx=(0, 5), pady=(0, 10))
            self.listbox1.bind('<<ListboxSelect>>', self.get_selected_items)
            self.listbox1.config(width=40)

            self.button_add = tk.Button(self.left_top_frame, text='>', command=self.add_item, width=2, height=2)
            self.button_add.pack(fill='both', expand=False, side='left', padx=(0, 5), pady=(0, 10))

            self.button_remove = tk.Button(self.left_top_frame, text='<', command=self.remove_item, width=2, height=2)
            self.button_remove.pack(fill='both', expand=False, side='left', padx=(0, 5), pady=(0, 10))

            self.listbox2 = tk.Listbox(self.left_top_frame, selectmode='single')
            self.listbox2.pack(fill='both', expand=True, side='left', padx=(0, 10), pady=(0, 10))
            self.listbox2.bind('<<ListboxSelect>>', self.get_selected_items)

            # Create the widgets in the left_bottom_frame



            # create a larger font for the button text
            button_font = font.Font(size=12)
            # create the button with the larger font
            self.save_as_button = tk.Button(self.left_bottom_frame, text='\U0001F4BE', font=button_font, command=self.saveas, width=3, height=1)

            self.save_as_button.pack(fill='both', expand=False, side='bottom', padx=(0, 5), pady=(0, 0), anchor='n')

            # Set the weights of the left_top_frame and left_bottom_frame to allocate space
            self.Output_left_frame.grid_columnconfigure(0, weight=1)
            self.Output_left_frame.grid_rowconfigure(0, weight=1)
            self.left_top_frame.grid_columnconfigure(0, weight=1)
            self.left_top_frame.grid_columnconfigure(1, weight=0)
            self.left_top_frame.grid_columnconfigure(2, weight=0)
            self.left_top_frame.grid_columnconfigure(3, weight=0)
            self.left_top_frame.grid_rowconfigure(0, weight=1)
            self.left_bottom_frame.grid_columnconfigure(0, weight=1)
            self.left_bottom_frame.grid_rowconfigure(0, weight=1)

            # drop row "Exact row source" and "Exact row target" from df_merged
            #self.df_merged = self.df_merged.drop(['Exact row source', 'Exact row target'], axis=1)
            
            for item in self.df_merged.columns.tolist():
                # add color to the listbox column from file1 and file2
                if item in self.df1.columns.tolist():
                    self.listbox1.insert(END, item)
                    self.listbox1.itemconfig(END, bg="Light Green")
                elif item.endswith('_x'):
                    self.listbox1.insert(END, item)
                    self.listbox1.itemconfig(END, bg="light Green")
                elif item.endswith('source'):
                    self.listbox1.insert(END, item)
                    self.listbox1.itemconfig(END, bg="light Green")
                elif item in self.df2.columns.tolist():
                    self.listbox1.insert(END, item)
                    self.listbox1.itemconfig(END, bg='#FFCCCB')
                elif item.endswith('_y'):
                    self.listbox1.insert(END, item)
                    self.listbox1.itemconfig(END, bg='#FFCCCB')
                elif item.endswith('target'):
                    self.listbox1.insert(END, item)
                    self.listbox1.itemconfig(END, bg='#FFCCCB')
                else:
                    self.listbox1.insert(tk.END, item)

            self.column_list = []
            for item in self.listbox2.get(0, 'end'):
                self.column_list.append(item)

            # merge df1 and df2 on column1 and column2 for Equal values
            self.df_merged = pd.merge(self.df1, self.df2, how='left', left_on=self.column1, right_on=self.column2)
            # remove nan
            self.df_merged = self.df_merged.fillna('')
            # show the merged dataframe in the treeview
            self.treeview = ttk.Treeview(self.Output_right_frame)

    def get_selected_items(self, event):
        self.listbox1_items = self.listbox1.curselection()
        self.listbox2_items = self.listbox2.curselection()

    # add color for item moved to listbox2 if _x in text bg='#e6dec3', if _ybg='#c3e4e6'
    def add_item(self):
        self.listbox2.delete(0, tk.END)
        for i in self.listbox1_items:
            self.listbox2.insert(END, self.listbox1.get(i))
            if self.listbox1.get(i).endswith('_x'):
                self.listbox2.itemconfig(END, bg="Light Green")
            elif self.listbox1.get(i).endswith('source'):
                self.listbox2.itemconfig(END, bg='light Green')
            # elif self.listbox1.get(i).endswith('Exact row source'):
            #     self.listbox2.itemconfig(END, bg='light Green')
            elif self.listbox1.get(i).endswith('_y'):
                self.listbox2.itemconfig(END, bg='#FFCCCB')
            elif self.listbox1.get(i).endswith('target'):
                self.listbox2.itemconfig(END, bg='#FFCCCB')                
            elif self.listbox1.get(i).endswith('Splitcolumn*'):
                self.listbox2.itemconfig(END, bg='#FFCCCB')
            # elif self.listbox1.get(i).endswith('Exact row target'):
            #     self.listbox2.itemconfig(END, bg='#FFCCCB')
            else:
                self.listbox2.itemconfig(END, bg='light Green')
        self.column_list = []

        # append item to the listbox2 in Output_treeview
        for item in self.listbox2.get(0, 'end'):
            self.column_list.append(item)

        merge_df = pd.merge(self.df1, self.df2, how='left', left_on=self.column1, right_on=self.column2)
        # remove nan
        merge_df = merge_df.fillna('')
        self.df_merged = merge_df[self.column_list]

    # if add item to listbox2, add column to Output_treeview
        self.column_list = []
        for item in self.listbox2.get(0, 'end'):
            self.column_list.append(item)

        self.Output_treeview['columns'] = self.column_list
        self.Output_treeview['show'] = 'headings'

        # mimimum width of columns in treeview is 100
        self.Output_treeview.column('#0', width=0, stretch=YES)


        for column in self.Output_treeview['columns']:
            self.Output_treeview.heading(column, text=column)
            self.Output_treeview.column(column, width=100, minwidth=100, stretch=YES)

        # populate treeview with data from listbox2
        for index, row in self.df_merged.iterrows():
            self.Output_treeview.insert('', index, values=list(row))

        self.Output_treeview.pack(fill='both', expand=True)

    def remove_item(self):
        self.listbox2_items = self.listbox2.curselection()
        for i in self.listbox2_items:
            self.listbox2.delete(i)

        # if remove item from listbox2, remove column from Output_treeview
        self.column_list = []
        for item in self.listbox2.get(0, 'end'):
            self.column_list.append(item)

        self.Output_treeview['columns'] = self.column_list
        self.Output_treeview['show'] = 'headings'
        for column in self.Output_treeview['columns']:
            self.Output_treeview.heading(column, text=column)
            self.Output_treeview.column(column, width=100, minwidth=100)

        # populate treeview with data from listbox2
        for index, row in self.df_merged.iterrows():
            self.Output_treeview.insert('', 'end', values=[row[column] for column in self.column_list])

    # create function to save merged columns in a new file
    def saveas(self):
        self.column_list = []

        # populate treeview with data from listbox2
        for item in self.listbox2.get(0, 'end'):
            self.column_list.append(item)

        self.file_name = filedialog.asksaveasfilename(
        filetypes=(("Excel file", "*.xlsx"), ("All files", "*.*")),
        defaultextension=".xlsx",
    )
        #self.df_merged
        self.df_merged.to_excel(self.file_name, index=False, columns=self.column_list)
        if messagebox.askyesno('Status', 'File Saved, do you want to open file?'):
            # start file with write permission
            try:
                os.startfile(os.path.normpath(self.file_name))
            except:
                messagebox.showerror("Error", "Cant open file, open manually")

    def search_in_listbox(self, event=None):

        # search_text_value in selected page in notebook tab. if run and one tab  is selected, search in that tab, so make the search_text_value universal to work on all pages when that page is selected.
        search_text_value = self.search_text.get()
                    
        # if selected_tab == 1:
            # if search_text_value is empty, clear selection and set bg color to white
        if search_text_value == '':
            for n, item in enumerate(self.Equal_listbox.get(0, tk.END)):
                self.Unequal_listbox.selection_clear(0, tk.END)
                self.Equal_listbox.itemconfig(n, bg='white')
            for n, item in enumerate(self.Unequal_listbox.get(0, tk.END)):
                self.Unequal_listbox.itemconfig(n, bg='white')
        elif search_text_value != '':
            # search in Equal listbox
            for n, item in enumerate(self.Equal_listbox.get(0, tk.END)):
                if search_text_value in str(item):
                    self.Equal_listbox.selection_clear(0, tk.END)
                    self.Equal_listbox.see(n)
                    self.Equal_listbox.itemconfig(n, bg='Light Green')
                else:
                    self.Equal_listbox.itemconfig(n, bg='white')
            # search in Unequal listbox
            for n, item in enumerate(self.Unequal_listbox.get(0, tk.END)):
                if search_text_value in str(item):
                    self.Unequal_listbox.selection_clear(0, tk.END)
                    self.Unequal_listbox.see(n)
                    self.Unequal_listbox.itemconfig(n, bg='#FFCCCB')
                else:
                    self.Unequal_listbox.itemconfig(n, bg='white')

    def reset_find_in_words_search(self):
        # Reset search-related variables to default values
        try:
            self.df3_temp = self.df2_temp
            self.tree.delete(*self.tree.get_children())
            self.progress_bar["value"] = 0
            self.max_value = len(self.source_column_data)
            self.cancel_button.pack_forget()
            self.cancel_button.grid_forget()
            self.cancel_button_pressed = False
        except:
            pass


    # find all matches from file1 and file2, including if it is just part of the word and not exact match
    def find_all_words(self):
        # if page with the same name is already open, delete it
        try:
            if self.notebook.tab(self.find_all_words_frame, option="text") == "Find all partial matches":
                # get index of tab "Find all partial matches"
                #tab_number = self.notebook.index(self.find_all_words_frame)
                self.notebook.forget(self.find_all_words_frame)
                
                self.find_all_words_frame = Frame(self.notebook)
                self.find_all_words_frame.pack(fill=BOTH, expand=True)
                self.notebook.add(self.find_all_words_frame, text="Find all partial matches")
                self.notebook.pack(fill="both", expand=True)
                # activate Find all partial matches frame in self.notebook
                self.notebook.select(self.find_all_words_frame)
        except:
            self.find_all_words_frame = Frame(self.notebook)
            self.find_all_words_frame.pack(fill=BOTH, expand=True)
            self.notebook.add(self.find_all_words_frame, text="Find all partial matches")
            self.notebook.pack(fill="both", expand=True)
            # activate Find all partial matches frame in self.notebook
            self.notebook.select(self.find_all_words_frame)

        # Reset search-related variables to default values
        self.reset_find_in_words_search()

        self.source_column = self.Source_column_combo.get()
        self.target_column = self.Target_column_combo.get()

        # get list of all values
        self.source_column_data = self.df1[self.source_column]
        self.target_column_data = self.df2[self.target_column]

        self.df2_temp = self.df2.copy()
        self.df2_temp[self.source_column] = ""

        # create a new empty column and insert values into it
        self.df2_temp[self.source_column] = ""

        # drop all column but target_column_data
        #self.df3_temp = self.df2_temp.drop(self.df2_temp.columns[[i for i in range(len(self.df2_temp.columns)) if i != self.df2_temp.columns.get_loc(self.target_column)]], axis=1)
        self.df3_temp = self.df2_temp.drop([col for col in self.df2_temp.columns if col != self.target_column], axis=1)

        self.new_column_data = self.df2_temp[self.source_column]

        # # create notebook page "Find all partial matches"
        # self.find_all_words_frame = Frame(self.notebook)
        # self.find_all_words_frame.pack(fill=BOTH, expand=True)

        self.notebook.add(self.find_all_words_frame, text="Find all partial matches")
        self.notebook.select(self.find_all_words_frame)

        with concurrent.futures.ThreadPoolExecutor(max_workers=4) as self.thread_pool:
            ...
        

        # add tree
        self.tree = ttk.Treeview(self.find_all_words_frame, columns=("", "Source", "Target"))
        self.tree.heading('#0', text='', anchor=W)
        self.tree.heading('#1', text='Source', anchor=W)
        self.tree.heading('#2', text='Target', anchor=W)
        self.tree.column('#0', stretch=NO, minwidth=10, width=0)
        self.tree.column('#1', stretch=YES, minwidth=300, width=380)
        self.tree.column('#2', stretch=YES, minwidth=600, width=700)

        # scroll bar for treeview horizontally and vertically
        self.scrollbar = tk.Scrollbar(self.find_all_words_frame, orient="vertical", command=self.tree.yview)
        self.scrollbar.pack(side=RIGHT, fill=Y)
        self.scrollbar2 = tk.Scrollbar(self.find_all_words_frame, orient="horizontal", command=self.tree.xview)
        self.scrollbar2.pack(side=BOTTOM, fill=X)
        self.tree.configure(yscrollcommand=self.scrollbar.set)
        self.tree.pack(fill=BOTH, expand=True, padx=5, pady=5)

        # new df
        self.df3_temp = self.df2_temp

        # # convert all text in source_column_data and target_column_data to string
        self.source_column_data = self.source_column_data.astype(str)
        self.target_column_data = self.target_column_data.astype(str)

        # show equal values in a window box
        self.Equal_values_label = tk.Label(self.master, text="Equal values:")

        # Create search text input
        self.search_text = tk.StringVar()
        self.search_text_entry = tk.Entry(self.find_all_words_frame, textvariable=self.search_text)
        self.search_text_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(20, 20), pady=(0, 0))
        self.search_text_entry.focus_set()
        self.search_text_entry.bind("<Return>", self.search_text_entry_on_enter)
            # create a larger font for the button text
        button_font = font.Font(size=12)
        # checkbutton to "Show not found values" to left side, and Save button, when save button is pressed give new popupmessage as to save with Original Source data, or save Treeview data
        self.Save_treeview_button = tk.Button(self.find_all_words_frame, text="\U0001F4BE Treeview",font=button_font, command=self.save_treeview)
        self.Save_treeview_button.pack(side=tk.LEFT, padx=(10, 0), pady=(0, 0))

        self.Save_with_Source_Button = tk.Button(self.find_all_words_frame, text="\U0001F4BE to Source data",font=button_font, command=self.Save_with_Source)
        self.Save_with_Source_Button.pack(side=tk.LEFT)

        # #add Save button at the end, when pressed ask user to save all Excel, or to save only Values in treeview
        self.Save_with_Target_Button = tk.Button(self.find_all_words_frame, text="\U0001F4BE to Target data",font=button_font, command=self.Save_with_Target)
        self.Save_with_Target_Button.pack(side=tk.LEFT)
        
        self.cancel_button = tk.Button(self.find_all_words_frame, text="Cancel", command=self.cancel_button_pressed, state=tk.NORMAL)
        self.cancel_button.pack(side=RIGHT, padx=(5, 20), pady=(0, 0))
        

        # update progress bar and cancel work if escape is pressed 
        self.progress_bar = ttk.Progressbar(self.find_all_words_frame, orient="horizontal", length=600, mode="determinate")
        self.progress_bar.pack(ipadx=50, padx=50)
        self.max_value = len(self.source_column_data)
        self.progress_bar["maximum"] = self.max_value
        self.progress_bar["value"] = 0
        self.progress_bar.update()
        self.find_all_words_frame.update()

        # hide save button show show after search is done
        self.Save_treeview_button.pack_forget()
        self.Save_with_Source_Button.pack_forget()
        self.Save_with_Target_Button.pack_forget()
       
        # find all matches and insert them into new column
        #start time
        self.start_time = time.time()
       # self.message_label = tk.Label(self.new2Window, text="Time used: " + str(round(time.time() - self.start_time, 2)) + " seconds, Time remaining: " + str(round(self.max_value - self.n, 2)) + " seconds")
        self.n = 0
        self.not_found_values = []
        for self.source_column_data_item in self.source_column_data:
            self.n += 1
            self.progress_bar["value"] = self.n
            self.progress_bar.update()
            self.find_all_words_frame.update()
            if self.cancel_button_pressed == True:
                break

            # Find all matches from self.target_column_data
            self.target_column_data_matches = self.target_column_data.str.findall(self.source_column_data_item, flags=re.I)

            # If there are matches, add them to the treeview
            if not all(matches == [] for matches in self.target_column_data_matches):
                for i, matches in enumerate(self.target_column_data_matches):
                    if matches:
                        self.tree.insert("", "end", text="", values=(self.source_column_data_item, self.target_column_data[i]))
            else:
                # Add not found values to treeview where value is not found
                new_row = pd.Series({'Column Name': self.source_column_data_item, 'Value': 'Not found'})
                self.new_column_data = self.new_column_data.append(new_row, ignore_index=True)
                self.not_found_values.append(self.source_column_data_item)
                self.tree.insert("", "end", text="", values=(self.source_column_data_item, "Not found"))
            

        self.progress_bar["value"] = self.max_value
        self.progress_bar.update()

        self.search_text_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(40, 40), pady=(0, 0))

        total_items = len(self.tree.get_children())
        time_used = round(time.time() - self.start_time, 2)
        self.message_label = tk.Label(self.find_all_words_frame, text=f"Finish: {total_items} Time used: {time_used} seconds")
        self.message_label.pack(side=tk.LEFT, padx=(20, 10), pady=(0, 0))
        self.message_label.update()

        # disable Cancel button
        self.cancel_button.pack_forget()
        # hide progress bar
        self.progress_bar.pack_forget()
        # show save button
        self.Save_treeview_button.pack(side=tk.LEFT)
        self.Save_with_Source_Button.pack(side=tk.LEFT)
        self.Save_with_Target_Button.pack(side=tk.LEFT)

    def selection_clear(self):
        self.tree.selection_clear(0, tk.END)

    def cancel_button_pressed(self):
        self.cancel_button_pressed = True
        # hide cancel button and show save button
        self.cancel_button.pack_forget()
        
        self.start_time = time.time()
        # Reset search-related variables to default values
        self.reset_find_in_words_search()
        #display label: Canceled + time
        self.message_label = tk.Label(self.find_all_words_frame, text="Canceled in " + str(int(time.time() - self.start_time) // 60) + " minutes and " + str(int(time.time() - self.start_time) % 60) + "s     ")
        self.progress_bar.pack_forget()
        self.Save_treeview_button.pack(side=tk.LEFT)
        self.cancel_button.pack(side=tk.LEFT)
        self.Save_with_Source_Button.pack(side=tk.LEFT)
        self.Save_with_Target_Button.pack(side=tk.LEFT)
        
        # Stop searching
        return


    def close_tab(self, event):

        # get name of the tab the user pressed the middle mouse button on
        # and delete it from the notebook but not for the first tab "Input files"

        tab_name = self.notebook.tab(self.notebook.select(), "text")
        if tab_name != "Input files":
            self.notebook.forget(self.notebook.select())

    # delete tree, update tree from df3
    def Save_with_Source(self):
        file_path = filedialog.asksaveasfilename(
            filetypes=(("Excel file", "*.xlsx"), ("All files", "*.*")),
            defaultextension=".xlsx",)

        # df1 = Source browse file
        self.df1 = pd.read_excel(self.Source_file_name, sheet_name=0, index_col=None, na_values=['NA'])
        
        children = self.tree.get_children()
        # create empty list
        values = []
        # loop through all children and get the values from column "Source" and "Target"
        for child in children:
            values.append(self.tree.item(child)["values"])
            
            #values.append([self.tree.item(child, "values")[0], self.tree.item(child, "values")[1]])
        # create dataframe from list
        df_treeview = pd.DataFrame(values)
        # rename columns
        df_treeview.columns = ["Source", "Target"]

        #Source_column_combo, merge Source with Source_column_combo.get()
        self.df1 = pd.merge(self.df1, df_treeview, how='left', left_on=self.Source_column_combo.get(), right_on='Source')
        # drop Source column
        self.df1 = self.df1.drop(columns=['Source'])
        # rename Target column to Source_column_combo.get()
        self.df1 = self.df1.rename(columns={'Target': self.Source_column_combo.get()})
        # save file
        # remove duplicates
        self.df1 = self.df1.drop_duplicates()
        self.df1.to_excel(file_path, index=False)

        if messagebox.askyesno('Status', 'File Saved, do you want to open file?'):
            # start file with write permission
            try:
                os.startfile(os.path.normpath(file_path))
            except:
                messagebox.showerror("Error", "Error saving file")
        else:
            pass

    def save_treeview(self):
        # get all children from treeview
        children = self.tree.get_children()
        # create empty list
        values = []
        # loop through all children and get the values from column "Source" and "Target"
        for child in children:
            values.append(self.tree.item(child)["values"])
            
            #values.append([self.tree.item(child, "values")[0], self.tree.item(child, "values")[1]])
        # create dataframe from list
        df_treeview = pd.DataFrame(values)
        # rename columns
        df_treeview.columns = ["Source", "Target"]

        file_path = filedialog.asksaveasfilename(
            filetypes=(("Excel file", "*.xlsx"), ("All files", "*.*")),
            defaultextension=".xlsx",
            )
        df_treeview.to_excel(file_path, index=False)
        # show messagebox "Do you want to open file?"
        if messagebox.askyesno('Status', 'File Saved, do you want to open file?'):
            # start file with write permission
            try:
                os.startfile(os.path.normpath(file_path))
            except:
                # if file is open
                messagebox.showerror('Error', 'File is open')
        else:
            pass

    # save all function
    def Save_with_Target(self):
        file_path = filedialog.asksaveasfilename(
            filetypes=(("Excel file", "*.xlsx"), ("All files", "*.*")),
            defaultextension=".xlsx",)


        # df3 = Source browse file
        self.df2 = pd.read_excel(self.Target_file_name, sheet_name=0, index_col=None, na_values=['NA'])
        # get all children from treeview
        children = self.tree.get_children()
        # create empty list
        values = []
        # loop through all children and get the values from column "Target" inn treeview
        for child in children:
            values.append(self.tree.item(child)["values"])
        # create dataframe from list
        df_treeview = pd.DataFrame(values)
        # rename columns
        df_treeview.columns = ["Source", "Target"]

        # for items in treevew column "Source", "Target" add to df2 where value in column "Target" is equal to value in df2 column self.Source_column_combo.get()
        for index, row in df_treeview.iterrows():
            try:
                self.df2.loc[self.df2[self.Source_column_combo.get()] == row["Source"], self.Target_column_combo.get()] = row["Target"]
            except:
                continue

        # at the end of df2 last row in column "Source" add self.not_found_values
        for item in self.not_found_values:
            new_row = pd.Series({self.Source_column_combo.get(): item, self.Target_column_combo.get(): 'Not found'})
            self.df2 = self.df2.append(new_row, ignore_index=True)


        # save df1 to excel file
        self.df2.to_excel(file_path, index=False)

        if messagebox.askyesno('Status', 'File Saved, do you want to open file?'):
            # start file with write permission
            try:
                os.startfile(os.path.normpath(file_path))
            except:
                messagebox.showerror("Error", "Error saving file")
        else:
            pass

    #search function
    def search_text_entry_on_enter(self, event):
        # not case sensitive search
        search_text = self.search_text.get()
        # search in treeview
        for item in self.tree.get_children():
            # if value in treeview contains search_text, color it red
            if search_text in str(self.tree.item(item)['values']):
                self.tree.item(item, tags=('Found',))
                # scroll to item
                self.tree.see(item)
            # if value in treeview does not contain search_text, color it white
            else:
                self.tree.item(item, tags=('Not found',))
        # if search_text is empty, color treeview white
        if self.search_text.get() == '':
            for item in self.tree.get_children():
                self.tree.item(item, tags=('Not found',))
        # tag configure
        self.tree.tag_configure('Found', background='#FFCCCB')
        self.tree.tag_configure('Not found', background='white')

    # def search_text_entry_on_enter(self, event):
    #     search_text = self.search_text.get()
    #     # clear tree
    #     for i in self.tree.get_children():
    #         self.tree.delete(i)

    #     # create a list of futures to track the threads
    #     self.futures = []

    #     # split the data into chunks and run search in parallel
    #     chunk_size = int(len(self.source_column_data) / 4)
    #     for i in range(0, len(self.source_column_data), chunk_size):
    #         start = i
    #         end = i + chunk_size
    #         self.futures.append(self.thread_pool.submit(self.search_matches, search_text, start, end))

    #     # wait for all threads to complete
    #     concurrent.futures.wait(self.futures)

    #     # update tree with matches
    #     self.update_tree()

    #     self.Save_treeview_button.pack(side=tk.LEFT)

    # def search_matches(self, search_text, start, end):
    #     for i in range(start, end):
    #         if search_text in self.source_column_data[i]:
    #             self.new_column_data[i] = self.source_column_data[i]

    # def update_tree(self):
    #     for i in range(len(self.new_column_data)):
    #         if self.new_column_data[i] != "":
    #             self.tree.insert("", tk.END, values=(self.new_column_data[i], self.target_column_data[i]))

    #     # join the threads
    #     for future in self.futures:
    #         future.result()
    def clear(self):
        # Delete all but the first tab
        for tab in self.notebook.tabs()[1:]:
            self.notebook.forget(tab)

        # Reset Source and Target file name labels
        self.Source_file_name_label.config(text="Choose file 1:")
        self.Target_file_name_label.config(text="Choose file 2:")

        # Reset Source and Target columns dropdown menus
        self.Source_column_combo['values'] = ()
        self.Source_column_combo.set("")
        self.Source_column_combo.config(state='disabled')

        self.Target_column_combo['values'] = ()
        self.Target_column_combo.set("")
        self.Target_column_combo.config(state='disabled')

        # Clear Source and Target treeviews
        self.Source_treeview["columns"] = ()
        self.Target_treeview["columns"] = ()
        self.Source_treeview.delete(*self.Source_treeview.get_children())
        self.Target_treeview.delete(*self.Target_treeview.get_children())

        # Clear merged dataframe and reset dataframes
        self.merged_dataframe = None
        self.Source_df = None
        self.Target_df = None

        self.equal_values = []
        self.unequal_values = []

        self.column1 = None
        self.column2 = None
        self.Source_file_name = None
        self.Target_file_name = None
        self.df1 = None
        self.df2 = None
        self.df1_column = None
        self.df2_column = None

        self.Search_Clear()


    def split_text(self):

        try:
            if self.notebook.tab(self.split_text_frame, option="text") == "Split text":
                # get index of tab "Split text"
                self.notebook.forget(self.split_text_frame)
                
                self.split_text_frame = Frame(self.notebook)
                self.split_text_frame.pack(fill=BOTH, expand=True)
                self.notebook.add(self.split_text_frame, text="Split text")
                self.notebook.pack(fill="both", expand=True)
                # activate Find all partial matches frame in self.notebook
                self.notebook.select(self.split_text_frame)

        except:
            self.split_text_frame = Frame(self.notebook)
            self.split_text_frame.pack(fill=BOTH, expand=True)
            self.notebook.add(self.split_text_frame, text="Split text")
            self.notebook.pack(fill="both", expand=True)
            # activate Find all "Split text" in self.notebook
            self.notebook.select(self.split_text_frame)

        self.Source_file_name_name = self.Source_file_name.split('/')[-1]
        #self.split_text_frame.configure(background="white")

        # df read source file
        self.df = pd.read_excel(self.Source_file_name)

        #rename space to _
        self.df.columns = self.df.columns.str.replace(' ', '_')

        # create frame for treeview
        self.frame = Frame(self.split_text_frame, background="light grey")
        self.frame.pack(fill=BOTH, expand=True)
        # create treeview
        self.tree2 = ttk.Treeview(self.frame, columns=self.df.columns.tolist(), show="headings")

        self.scrollbar = Scrollbar(self.frame, orient="vertical", command=self.tree2.yview)
        self.scrollbar.pack(side=RIGHT, fill=Y)
        self.scrollbar2 = Scrollbar(self.frame, orient="horizontal", command=self.tree2.xview)
        self.scrollbar2.pack(side=BOTTOM, fill=X)

        # set scrollbar to treeview
        self.tree2.configure(yscrollcommand=self.scrollbar.set)
        self.tree2.configure(xscrollcommand=self.scrollbar2.set)

        self.tree2.pack(fill=BOTH, expand=True)
        self.tree2.heading('#0', text='Index')
        self.tree2.column('#0', anchor='center', stretch=YES, minwidth=10, width=50)

        # Add text input and ask for character to split text on, and input box and column dropdown menu in the same row
        self.input_field = Entry(self.split_text_frame)
        self.input_field.insert(0, "Character to search for:")
        # when entered clear text
        self.input_field.pack(side=LEFT, fill=X, expand=True, padx=(10, 5), pady=(10, 0))
        self.input_field.bind("<Button-1>", self.clear_text)
        self.input_field.config(state=DISABLED)

        # Add text input and ask for character to split text on, and input box and column dropdown menu in the same row
        self.input2_field = Entry(self.split_text_frame)
        self.input2_field.insert(0, "Character to replace with:")
        # when entered clear text
        self.input2_field.pack(side=LEFT, fill=X, expand=True, padx=(10, 5), pady=(10, 0))

        self.columns_dropdown = Combobox(self.split_text_frame, values=self.df.columns.tolist())
        self.columns_dropdown.pack(side=LEFT, fill=X, expand=True, padx=(0, 10), pady=(10, 0))
        # event when change get colum
        self.columns_dropdown.bind("<<ComboboxSelected>>", self.get_column)
        self.input2_field.config(state=DISABLED)

        # add button to find and replace text
        self.findreplace_button = Button(self.split_text_frame, text="Find and Replace", command=self.find_text)
        self.findreplace_button.pack(side=LEFT, fill=X, expand=True, padx=(0, 10), pady=(10, 0))

        # Add a button to execute function split_words
        self.split_text_button = tk.Button(self.split_text_frame, text="Split word", width=12, command=self.split_words)
        self.split_text_button.pack(side=LEFT, fill=X, expand=True, padx=(0, 5), pady=(10, 0))
        # disable
        # create a larger font for the button text
        button_font = font.Font(size=12)
        # Add a button to save splitted word file
        self.Save_file_button = tk.Button(self.split_text_frame, text="\U0001F4BE",font=button_font, width=12, command=self.save_as)
        self.Save_file_button.pack(side=RIGHT, fill=X, expand=True, padx=(5, 10), pady=(10, 0))
        self.Save_file_button.config(state='disabled')

        # create list with columns from dataframe
        self.columns = self.df.columns.tolist()
        self.columns_dropdown['values'] = self.columns
        self.columns_dropdown.current(0)
        self.columns_dropdown.config(state='readonly')

        self.df = self.df.replace(np.nan, '', regex=True)

        # header has no index
        self.tree2.heading('#0', text='Index')
        self.tree2.column('#0', anchor='center', stretch=YES, minwidth=10, width=50)

        # # add column headers to treeview
        for i in self.columns:
            self.tree2.heading(i, text=i)
            self.tree2.column(i, anchor='center', stretch=YES, minwidth=10, width=50)
        
        # add data to treeview
        for i in range(len(self.df)):
            self.tree2.insert('', 'end', values=self.df.iloc[i].tolist())

    # update treeview with new column
    def get_column(self, event):
        self.input_field.config(state=NORMAL)
        # update treeview with new column
        self.column = self.columns_dropdown.get()
        
        self.tree2.delete(*self.tree2.get_children())
        self.tree2.configure(columns="")
        self.tree2.delete(*self.tree2.get_children())

        # create a new tree with new column
        self.tree2.configure(columns=self.column) ########## WIDTH!!!
        
        self.tree2.heading(self.column, text=self.column)
        self.tree2.column(self.column, anchor='center', stretch=YES, minwidth=200, width=500)

        # get column data
        self.column_data = self.df[self.column]

        # insert column data into treeview
        for i in range(len(self.column_data)):
            self.tree2.insert("", "end", text=i, values=(self.column_data[i], ""))
        
        self.input2_field.config(state=NORMAL)

    def split_words(self):
        # new df copy of selected column
        self.column_data = self.df[self.column]

        # split text on character and display in new df
        self.column_data = self.column_data.str.split(self.input_field.get())
        self.column_data = pd.DataFrame(self.column_data.tolist())
        self.column_data = self.column_data.replace(np.nan, '', regex=True)

        # delete old tree, remove all headers, and all columns
        self.tree2.delete(*self.tree2.get_children())
        self.tree2.configure(columns="")

        # add header in self.df1 name Column 1, Column 2, Column 3, ...
        self.column_data.columns = ["Splitcolumn " + str(i) for i in range(1, len(self.column_data.columns)+1)]

        # create a new tree with new column
        self.tree2.configure(columns=self.column_data.columns.tolist())
        # header has no index
        self.tree2.heading('#0', text='Index')
        self.tree2.column('#0', anchor='center', stretch=YES, minwidth=10, width=50)

        # # add column headers to treeview
        for i in range(len(self.column_data.columns)):
            self.tree2.heading(i, text=self.column_data.columns[i])
            self.tree2.column(i, anchor='center', stretch=YES, minwidth=10, width=100)

        # add dataframe to treeview ( in index keep the word as originale before the split)
        for i in range(len(self.column_data.index)):
            self.tree2.insert('', 'end', text=str(self.column_data.index[i]), values=tuple(self.column_data.iloc[i])) 
        self.Save_file_button.config(state='normal')

    def find_text(self):
    # ask what to search for and what to replace with, use self.input_field.
    # find values from user and replace old with new, allow wildcard search "*"
        self.split_text_button.config(state='normal')      

        self.find_texts = self.input_field.get()
        self.replace_text = self.input2_field.get()
        # enable
        self.column_data = self.df[self.column]

        # replace text in column
        self.column_data = self.df[self.column].replace(self.find_texts, self.replace_text, regex=True)

        # delete old tree, remove all headers, and all columns
        self.tree2.delete(*self.tree2.get_children())
        self.tree2.configure(columns="")
        
        # add changes to df and view df in tree
        self.df[self.column] = self.column_data
        self.tree2.delete(*self.tree2.get_children())
        self.tree2.configure(columns="")
        self.tree2.delete(*self.tree2.get_children())

        # create a new tree with new column
        self.tree2.configure(columns=self.column) ########## WIDTH!!!
        # header name same as dropdown menu
        
        self.tree2.heading(self.column, text=self.column)
        self.tree2.column(self.column, anchor='center', stretch=YES, minwidth=200, width=500)

        # get column data
        self.column_data = self.df[self.column]

        # insert column data into treeview
        for i in range(len(self.column_data)):
            self.tree2.insert("", "end", text=i, values=(self.column_data[i], ""))

        self.Save_file_button.config(state='normal')

    def clear_text(self, event):
        # disable input2_field with message "enter text to search for"
        self.input_field.delete(0, END)
        self.input2_field.delete(0, END)

    def save_as(self):
        #save without index
        self.splitdf = self.column_data.copy()
        file_path = filedialog.asksaveasfilename(
        filetypes=(("Excel file", "*.xlsx"), ("All files", "*.*")),
        defaultextension=".xlsx",
    )

        # add back to same file with new columns from the split
        self.splitdf = pd.concat([self.df, self.splitdf], axis=1)
        self.splitdf.to_excel(file_path, index=False)


        #self.splitdf.to_excel(file_path, index=False)
        if messagebox.askyesno('Status', 'File Saved, do you want to open file?'):
            # start file with write permission
            try:
                os.startfile(os.path.normpath(file_path))
            except:
                messagebox.showerror("Error", "Cant open file, open manually")


    ##############################################################################################
    ####################################  SEARCH IN FOLDER #######################################
    ##############################################################################################
    def select_folder(self):

        # select folder to search in
        self.folder_selected = filedialog.askdirectory(initialdir="/", title="Select folder")
        # update label with folder name
        self.label_select_folder.configure(text=self.folder_selected)
        # pass if cancel is pressed
        if self.folder_selected == "":
            pass

        #Enable self.Search_btn1
        self.Search_btn1.config(state='normal')

    def Search_OnEnter(self, event):
        # red text on mouseover
        self.Search_checkbutton.configure(foreground="red")
        self.Search_checkbutton.configure(text="Search in subfolders might take a while depending on the number of files in the folder")
        
    def Search_OnLeave(self, event):
        self.Search_checkbutton.configure(foreground="black")
        self.Search_checkbutton.configure(text="Search in subfolders")     

    def Search_SearchCSV(self, file, value):
        # Initialize the stop flag
        self.stop_flag = False
        matches = []
        try:
            with open(file, newline='') as csvfile:
                reader = csv.reader(csvfile)
                header = next(reader)
                for i, row in enumerate(reader):
                    # Check the stop flag
                    if self.stop_flag:
                        return matches
                    for j, cell in enumerate(row):
                        if (self.Search_box2_value.get() == 'Exact match' and cell == value) or (value in str(cell)):
                            self.Search_tree.bind("<Escape>", self.Search_stop)
                            # Add 1 to i and j to convert to 1-based indexing
                            matches.append((i+1, j+1))
                            self.Search_tree.insert('', 'end', text=str(cell), values=(file, j+1, i+1, cell))
        except:
            pass
        return matches


    #### ÅPNE EXCEL FIL OG SØK I DEN ####
    def Search_OnDoubleClick(self, event):
        item = self.Search_tree.selection()[0]
        file_path = self.Search_tree.item(item, "values")[0]
        cell = self.Search_tree.item(item, "values")[2]

        # if end with .xlsx
        if file_path.endswith(".xlsx"):
            # if on windows
            if os.name == 'nt':
                # Use win32com.client to access the Excel application through COM
                excel = win32com.client.Dispatch("Excel.Application")
                # Use the Workbooks.Open method to open the Excel file
                wb = excel.Workbooks.Open(file_path)
                # Use the Visible property to make the Excel application visibl
                excel.Visible = True
                # Use the Sheets property to access the Excel worksheet
                ws = wb.Sheets(1)
                # Use the Range method to access the cell
                cell = ws.Range(cell)
                # Use the Select method to select the cell
                cell.Select()
                # Use the Activate method to activate the cell
                cell.Activate()
            
            # if on macbook
            elif os.name == 'posix':
                # Use the subprocess module to open the Excel file
                subprocess.run(["open", "-a", "Microsoft Excel", file_path])

        self.Search_tree.selection_set(self.Search_tree.identify_row(event.y))    
        item = self.Search_tree.selection()[0]
        file = self.Search_tree.item(item, "values")[0]
        # if list is empty, don't do anything
        if self.Search_tree.get_children().__len__() == 0:
            return




        # TODO: Scroller ikke helt riktig når jeg åpner csv. Men ellers så fungerer resten bra.
        if file_path.lower().endswith((".csv")):
            excel = win32com.client.Dispatch("Excel.Application")
            wb = excel.Workbooks.Open(file_path)
            excel.Visible = True
            ws = wb.Sheets(1)
            
            # Convert 1-based row and column indexes to 0-based indexes
            row_index = cell - 1
            col_index = 0
            
            # Check that the row and column indexes are within the bounds of the worksheet
            if row_index >= ws.Rows.Count or col_index >= ws.Columns.Count:
                messagebox.showerror("Error", "Row or column index out of range")
            else:
                cell_address = ws.Cells(row_index+1, col_index+1)
                cell_address.Select()
                cell_address.Activate()


        if self.Search_tree.get_children().__len__() == 0:
            return

        if file.endswith('.pdf'):
            # def open_pdf(self, event):
            item = self.Search_tree.selection()[0]
            file = self.Search_tree.item(item, "values")[0]
            os.startfile(file)    
            # go to page in pdf

        # if found in word, display document in tree with page value is found, and value            
        if file.endswith('.docx'):
            # def open_word(self, event):
            item = self.Search_tree.selection()[0]
            file = self.Search_tree.item(item, "values")[0]
            os.startfile(file)
        
    def Search_Clear(self):

        self.Search_tree.delete(*self.Search_tree.get_children())
        self.Search_entry1.delete(0, 'end')
        self.Search_status['text'] = "No folder selected"
        self.Search_progress['value'] = 0
        self.Search_progress.stop()
        self.label_select_folder.configure(text="No folder selected", background="#FFCCCB", foreground="black")
        # disable Search button
        self.Search_btn1.config(state="disabled")
        #unclick checkbox
        self.Search_checkbutton.deselect()
        
    def Search_exitwindow(self, event):
        #hide window
        self.Search_tree.bind("<Leave>", self.Search_exitwindow)
        # wait 10 seconds and then withdraw the window
        self.Search_window.after(5000, self.Search_window.withdraw)

   # open files and navigate to the cell number in that file from treeview    
    def Search_open_file(self, event):
        item = self.Search_tree.selection()[0]
        file = self.Search_tree.item(item, "values")[0]
        os.startfile(file)



    def Search_OnButtonClick(self):
        # if Search_entry1 is empty, show error message "Enter search value first"
        if self.Search_entry1.get() == "":
            messagebox.showerror("Error", "Enter search value first")
            return

        # if select_folder has something else than "Select folder to search in", then continue else show error message "Select folder first
        if self.label_select_folder.cget("text") != "No folder selected":
            self.Search_tree.delete(*self.Search_tree.get_children())
            self.Search_path = self.folder_selected
            self.Search_value = self.Search_entry1.get()
            self.Search_files = self.Search_FindFiles(self.Search_path, self.Search_value)
            self.Search_progress['maximum'] = len(self.Search_files)
            self.Search_progress['value'] = 0
            self.Search_status['text'] = "Searching...   "
            self.search_cancel = False
            self.master.bind("<Escape>", self.Search_stop) # bind escape key to Search_stop function

            start_time = time.time() # start timer

            for f in self.Search_files:
                if self.search_cancel:  # check if search_cancel flag is set to True
                    break  # exit the for loop
                self.Search_progress['value'] += 1
                self.Search_progress.update()
                self.Search_status['text'] = "Searching...    {0}".format(f)
                self.Search_status.update()
                self.Search_SearchFile(f, self.Search_value)

            # end timer and calculate elapsed time
            elapsed_time = time.time() - start_time

            # self status show how many files, how many matches and how long the search took
            self.Search_status['text'] = "Search completed: Found {0} matches in {1} files. It took {2:.2f} seconds.".format(
                self.Search_tree.get_children().__len__(), len(self.Search_files), elapsed_time)

            self.master.unbind("<Escape>") # unbind escape key
            self.Search_progress.stop()
        else:
            messagebox.showerror("Error", "Select folder first")


    def Search_single_click(self, event):
        self.Search_tree.selection_set(self.Search_tree.identify_row(event.y))    
        item = self.Search_tree.selection()[0]
        file = self.Search_tree.item(item, "values")[0]
        # if list is empty, don't do anything
        if self.Search_tree.get_children().__len__() == 0:
            return
        else:           
            if file.endswith((".xlsx", ".xls", ".xlsm")):
                # select the item the mouse is over
                self.Search_tree.selection_set(self.Search_tree.identify_row(event.y))    
                # df drop na
                # df = pd.read_excel(self.Search_tree.item(self.Search_tree.selection()[0], "values")[0])
                df = pd.read_excel(self.Search_tree.item(self.Search_tree.selection()[0], "values")[0])
                df = df.dropna(axis=1, how='all')
                df = df.dropna(axis=0, how='all')
                
                # Open new window with pandas dataframe
                self.Search_window = Toplevel(self.master)
                # title of new window is the file name
                self.Search_window.title(self.Search_tree.item(self.Search_tree.selection()[0], "values")[0])
                self.Search_window.geometry("1000x600")
                self.Search_window.resizable(True, True)
                self.Search_window.configure(background="white")

                # create frame for treeview
                self.Search_frame = Frame(self.Search_window, background="white")
                self.Search_frame.pack(fill=BOTH, expand=True)
                # create treeview
                self.Search_tree2 = ttk.Treeview(self.Search_frame, columns=df.columns, show="headings")

                self.Search_scrollbar = ttk.Scrollbar(self.Search_frame, orient="vertical", command=self.Search_tree2.yview)
                self.Search_scrollbar.pack(side=RIGHT, fill=Y)
                self.Search_scrollbar2 = ttk.Scrollbar(self.Search_frame, orient="horizontal", command=self.Search_tree2.xview)
                self.Search_scrollbar2.pack(side=BOTTOM, fill=X)

                # set scrollbar to treeview
                self.Search_tree2.configure(yscrollcommand=self.Search_scrollbar.set)
                self.Search_tree2.configure(xscrollcommand=self.Search_scrollbar2.set)

                self.Search_tree2.pack(fill=BOTH, expand=True)
                self.Search_tree2.heading('#0', text='Index')
                self.Search_tree2.column('#0', anchor='center', stretch=YES, minwidth=10, width=50)

                # remove nan and na in df
                df = df.fillna('')
                df = df.replace(np.nan, '', regex=True)
                
                for i in range(len(df.columns)):
                    self.Search_tree2.heading(i, text=df.columns[i])
                    self.Search_tree2.column(i, anchor='center', width=100)
                # add dataframe to treeview
                for i in range(len(df.index)):
                    self.Search_tree2.insert('', 'end', text=str(df.index[i]), values=tuple(df.iloc[i]))

                cell = int(self.Search_tree.item(self.Search_tree.selection()[0], "values")[2][1:])
                # split self.Search_tree.item(self.Search_tree.selection()[0], "values")[2]) to get Column and Row, letter is column and number is row
                column = self.Search_tree.item(self.Search_tree.selection()[0], "values")[2][0]
                # select cell
                self.Search_tree2.selection_set(self.Search_tree2.get_children()[cell - 2])
                # scroll to cell
                if self.Search_tree.item(self.Search_tree.selection()[0], "values")[2][1:].isnumeric():
                    cell = int(self.Search_tree.item(self.Search_tree.selection()[0], "values")[2][1:])
                else:
                    pass
                try:
                    row = self.Search_tree2.get_children()[cell - 2]
                    self.Search_tree2.see(row)
                    self.Search_tree2.focus(row)
                    self.Search_tree2.selection_set(row)
                except ValueError:
                    pass



        if file.lower().endswith(".csv"):
            with open(self.Search_tree.item(self.Search_tree.selection()[0], "values")[0], 'r', encoding='utf-8-sig') as f:
                df = pd.read_csv(f)

                # Open new window with pandas dataframe
                self.Search_window = Toplevel(self.master)
                # title of new window is the file name
                self.Search_window.title(self.Search_tree.item(self.Search_tree.selection()[0], "values")[0])
                self.Search_window.geometry("1000x600")
                self.Search_window.resizable(True, True)
                self.Search_window.configure(background="white")

                # create frame for treeview
                self.Search_frame = Frame(self.Search_window, background="white")
                self.Search_frame.pack(fill=BOTH, expand=True)
                # create treeview
                self.Search_tree2 = ttk.Treeview(self.Search_frame, columns=df.columns, show="headings")

                self.Search_scrollbar = ttk.Scrollbar(self.Search_frame, orient="vertical", command=self.Search_tree2.yview)
                self.Search_scrollbar.pack(side=RIGHT, fill=Y)
                self.Search_scrollbar2 = ttk.Scrollbar(self.Search_frame, orient="horizontal", command=self.Search_tree2.xview)
                self.Search_scrollbar2.pack(side=BOTTOM, fill=X)

                # set scrollbar to treeview
                self.Search_tree2.configure(yscrollcommand=self.Search_scrollbar.set)
                self.Search_tree2.configure(xscrollcommand=self.Search_scrollbar2.set)

                self.Search_tree2.pack(fill=BOTH, expand=True)
                self.Search_tree2.heading('#0', text='Index')
                self.Search_tree2.column('#0', anchor='center', stretch=YES, minwidth=10, width=50)

                for i in range(len(df.columns)):
                    self.Search_tree2.heading(i, text=df.columns[i])
                    self.Search_tree2.column(i, anchor='center', width=100)

                for i in range(len(df.index)):
                    self.Search_tree2.insert('', 'end', text=str(df.index[i]), values=tuple(df.iloc[i]))

            # scroll to row where value is found
            search_value = self.Search_tree.item(self.Search_tree.selection()[0], "values")[3]
            try:
                row_index = df.index[df.eq(search_value).any(1)].tolist()[0]
                row = self.Search_tree2.get_children()[row_index]
                self.Search_tree2.see(row)
                self.Search_tree2.focus(row)
                self.Search_tree2.selection_set(row)
            except IndexError:
                pass

            # if found in pdf, display document in tree with page value is found, and value
            if file.endswith('.pdf'):
                # def open_pdf(self, event):
                item = self.Search_tree.selection()[0]
                file = self.Search_tree.item(item, "values")[0]
                os.startfile(file)    
                # go to page in pdf

            # if found in word, display document in tree with page value is found, and value            
            if file.endswith('.docx'):
                # def open_word(self, event):
                item = self.Search_tree.selection()[0]
                file = self.Search_tree.item(item, "values")[0]
                os.startfile(file)

    # if checkbutton is pressed search for in all subfolder in selected folder
    def Search_FindFiles(self, path, value):
        if self.Search_checkbutton_value.get() == 1:
            files = []
            for root, dirnames, filenames in os.walk(path):
                for filename in fnmatch.filter(filenames, '*.*'):
                    files.append(os.path.join(root, filename))
            return files
        else:
            return glob.glob(os.path.join(path, '*.*'))

    def Search_SearchFile(self, file, value):
        if self.Search_box_value.get() == 'Excel' or self.Search_box_value.get() == 'Excel, CSV, Word and PDF':
            self.Search_SearchExcel(file, value)
        if self.Search_box_value.get() == 'CSV' or self.Search_box_value.get() == 'Excel, CSV, Word and PDF':
            self.Search_SearchCSV(file, value)
        if self.Search_box_value.get() == 'Excel, CSV, Word and PDF':
            self.Search_SearchExcel(file, value)
            self.Search_SearchCSV(file, value)
        if self.Search_box_value.get() == 'PDF' or self.Search_box_value.get() == 'Excel, CSV, Word and PDF':
            self.Search_SearchPDF(file, value)
        if self.Search_box_value.get() == 'Word' or self.Search_box_value.get() == 'Excel, CSV, Word and PDF':
            self.Search_SearchWord(file, value)

    # def Search_SearchExcel(self, file, value):
    #     # Initialize the stop flag
    #     self.stop_flag = False
    #     try:
    #         wb = load_workbook(file)
    #         self.Search_tree.bind("<Escape>", self.Search_stop)
    #         for sheet in wb:
    #             # Check the stop flag
    #             if self.stop_flag:
    #                 return
    #             self.Search_tree.bind("<Escape>", self.Search_stop)
    #             for row in sheet.iter_rows():
    #                 # Check the stop flag
    #                 if self.stop_flag:
    #                     return
    #                 cells = (cell for cell in row if cell.value is not None)
    #                 matches = (cell for cell in cells if (self.Search_box2_value.get() == 'Exact match' and cell.value == value) or (value in str(cell.value)))
    #                 for match in matches:
    #                     self.Search_tree.bind("<Escape>", self.Search_stop)
    #                     self.Search_tree.insert('', 'end', text=str(match.value), values=(file, sheet.title, match.coordinate, match.value))
                    
    #     except:
    #         pass  
    
    # def Search_SearchExcel(self, file, value):
    #     # Initialize the stop flag
    #     self.stop_flag = False
    #     try:
    #         wb = load_workbook(file)
    #         self.Search_tree.bind("<Escape>", self.Search_stop)
    #         for sheet in wb:
    #             # Check the stop flag
    #             if self.stop_flag:
    #                 return
    #             self.Search_tree.bind("<Escape>", self.Search_stop)
    #             for row in sheet.iter_rows():
    #                 # Check the stop flag
    #                 if self.stop_flag:
    #                     return
    #                 cells = (cell for cell in row if cell.value is not None)
    #                 matches = (cell for cell in cells if (self.Search_box2_value.get() == 'Exact match' and cell.value == value) or (value in str(cell.value)))
    #                 for match in matches:
    #                     self.Search_tree.bind("<Escape>", self.Search_stop)
    #                     self.Search_tree.insert('', 'end', text=str(match.value), values=(file, sheet.title, match.coordinate, match.value))
                    
    #     except:
    #         pass

    def Search_SearchExcel(self, file, value):
        # Initialize the stop flag
        self.stop_flag = False
        try:
            wb = load_workbook(file)
            self.Search_tree.bind("<Escape>", self.Search_stop)
            for sheet in wb:
                # Check the stop flag
                if self.stop_flag:
                    return
                self.Search_tree.bind("<Escape>", self.Search_stop)
                for row in sheet.iter_rows():
                    # Check the stop flag
                    if self.stop_flag:
                        return
                    cells = (cell for cell in row if cell.value is not None)
                    if self.Search_box2_value.get() == 'Exact match':
                        matches = (cell for cell in cells if cell.value == value)
                    elif self.Search_box2_value.get() == 'Partial match':
                        matches = (cell for cell in cells if value in str(cell.value))
                    elif self.Search_box2_value.get() == 'Fuzzy match':
                        matches = (cell for cell in cells if fuzz.token_sort_ratio(str(cell.value), value) >= 80)
                    else:
                        # Default to partial match search
                        matches = (cell for cell in cells if value in str(cell.value))
                    for match in matches:
                        self.Search_tree.bind("<Escape>", self.Search_stop)
                        self.Search_tree.insert('', 'end', text=str(match.value), values=(file, sheet.title, match.coordinate, match.value))
        except:
            pass



    # def Search_SearchPDF(self, file, value):
    #     try:
    #         pdfFileObj = open(file, 'rb')
    #         pdfReader = PyPDF2.PdfFileReader(pdfFileObj)
    #         pages = (pdfReader.getPage(page) for page in range(pdfReader.numPages))
    #         pattern = re.compile(re.escape(value), re.IGNORECASE)
    #         for i, page in enumerate(pages):
    #             matches = pattern.findall(page.extractText())
    #             if matches:
    #                 for match in matches:
    #                     self.Search_tree.insert('', 'end', text=str(match), values=(file, '{0}'.format(i + 1), '', match))
    #         pdfFileObj.close()
    #     except:
    #         pass

    def Search_SearchPDF(self, file, value):
        try:
            pdfFileObj = open(file, 'rb')
            pdfReader = PyPDF2.PdfFileReader(pdfFileObj)
            for page in range(pdfReader.numPages):
                pageObj = pdfReader.getPage(page)
                if self.Search_box2_value.get() == 'Exact match':
                    if value in pageObj.extractText():
                        self.Search_tree.insert('', 'end', text=str(value), values=(file, '{0}'.format(page + 1), '', value))
                else:
                    if value in pageObj.extractText():
                        self.Search_tree.insert('', 'end', text=str(value), values=(file, '{0}'.format(page + 1), '', value))
        except:
            pass

    # def Search_SearchWord(self, file, value):
    #     try:
    #         if os.path.splitext(file)[1] == '.docx':
    #             doc = docx2txt.process(file)
    #             if self.Search_box2_value.get() == 'Exact match':
    #                 if value in doc:
    #                     self.Search_tree.insert('', 'end', text=str(value), values=(file, '', '', value))
    #             else:
    #                 if re.search(value, doc):
    #                     self.Search_tree.insert('', 'end', text=str(value), values=(file, '', '', value))
    #     except:
    #         pass
    def Search_SearchWord(self, file, value):
        try:
            doc = docx2txt.process(file)
            if self.Search_box2_value.get() == 'Exact match':
                # exact value match
                pattern = r'\b{}\b'.format(re.escape(value))
                if re.search(pattern, doc):
                    self.Search_tree.insert('', 'end', text=str(value), values=(file, '', '', value))
            elif self.Search_box2_value.get() == 'Partial match':
                # partial value match
                pattern = r'{}'.format(re.escape(value))
                for match in re.findall(pattern, doc):
                    self.Search_tree.insert('', 'end', text=str(match), values=(file, '', '', match))
            else:
                # fuzzy match
                matches = process.extractBests(value, [doc], scorer=fuzz.token_sort_ratio, score_cutoff=80)
                for match in matches:
                    self.Search_tree.insert('', 'end', text=str(match[0]), values=(file, '', '', match[0]))
        except:
            pass

    def Search_stop(self, event):
        self.master.unbind("<Escape>")
        self.search_cancel = True  # set the search_cancel flag to True
        answer = messagebox.askyesno("Stop search", "Are you sure you want to stop the search?")
        # if yes, break the loop and cancel the search
        if answer:
            self.search_cancel = True  # set the search_cancel flag to True
            self.Search_progress.stop()
            self.Search_status['text'] = "Search cancelled: Found {0} matches in {1} files. it took {2}s".format(self.Search_tree.get_children().__len__(), len(self.Search_files))

    def Search_Save(self):
        file_path = filedialog.asksaveasfilename(
            filetypes=(("Excel file", "*.xlsx"), ("All files", "*.*")),
            defaultextension=".xlsx",
        )

        if not file_path:
            # User cancelled the save operation
            return

        # Change the cursor to a wait cursor
        self.master.config(cursor="wait")

        # Get the values from the treeview and store them in a list of lists
        search_results = [self.Search_tree.item(i, "values") for i in self.Search_tree.get_children()]

        if not search_results:
            # No search results found, do not save anything
            messagebox.showinfo('Status', 'No search results found.')
            return

        # Create a dataframe from the list of lists
        df = pd.DataFrame(search_results, columns=['File', 'Sheet', 'Cell', 'Value'])

        # Save the dataframe to an excel file
        df.to_excel(file_path, index=False)

        # Change the cursor back to the default cursor
        self.master.config(cursor="")

        # Show a message box indicating the file was saved
        messagebox.showinfo('Status', 'File saved to: {}'.format(file_path))

        # Open the file if the user wants to
        if messagebox.askyesno('Status', 'Do you want to open the file?'):
            try:
                os.startfile(os.path.normpath(file_path))
            except:
                messagebox.showinfo("Error", "File not found")

def main():
    root = tk.Tk()
    app = MainWindow(root)
    root.title('Search & Index')
    root.geometry('1400x800')
    #root.wm_iconbitmap(icon_path)
    #root.iconbitmap(r"Icon/Textinfolder.png")
    root.mainloop()

if __name__ == '__main__':
    main()
